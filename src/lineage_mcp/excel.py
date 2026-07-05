"""Render a LineageGraph as a clean source-to-target mapping (STTM) workbook.

Up to three sheets:
  - "Flow Diagram": a rendered boxes-and-arrows flowchart image of the
    table-level lineage (only added if Graphviz's `dot` executable is
    available - see diagram.py).
  - "Table Lineage": one row per source-table -> target-table relationship.
  - "Column Lineage": one row per source-column -> target-column relationship
    (only added if the graph actually has column-level edges - e.g. SQL or
    Informatica parses; SSIS is component-level only and won't have one).

`direction` controls which entity is presented first (left-to-right) to
match how the mapping reads ("source to target" vs "target to source");
the underlying values are always the true source/target, never swapped.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from lineage_mcp.diagram import render_flow_diagram_png
from lineage_mcp.graph import LineageGraph, NodeKind

_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def to_excel_workbook(full_graph: LineageGraph, direction: str = "source_to_target") -> Workbook:
    if direction not in ("source_to_target", "target_to_source"):
        raise ValueError("direction must be 'source_to_target' or 'target_to_source'")

    table_graph = full_graph.collapse_to_kind(NodeKind.TABLE)

    wb = Workbook()
    diagram_png = render_flow_diagram_png(table_graph, direction=direction) if not table_graph.is_empty() else None
    if diagram_png:
        diagram_sheet = wb.active
        diagram_sheet.title = "Flow Diagram"
        _write_diagram_sheet(diagram_sheet, diagram_png)
        table_sheet = wb.create_sheet("Table Lineage")
    else:
        table_sheet = wb.active
        table_sheet.title = "Table Lineage"

    _write_table_sheet(table_sheet, table_graph, direction)

    column_rows = _column_rows(full_graph)
    if column_rows:
        column_sheet = wb.create_sheet("Column Lineage")
        _write_column_sheet(column_sheet, column_rows, direction)

    return wb


def _write_diagram_sheet(ws, png_bytes: bytes) -> None:
    ws.add_image(Image(io.BytesIO(png_bytes)), "A1")
    ws.sheet_view.showGridLines = False


def _write_table_sheet(ws, table_graph: LineageGraph, direction: str) -> None:
    first_header, second_header = ("Source Table", "Target Table") if direction == "source_to_target" else ("Target Table", "Source Table")
    headers = [first_header, second_header, "Detail"]

    rows = []
    for e in table_graph.edges:
        src_label = table_graph.nodes[e.source].label
        tgt_label = table_graph.nodes[e.target].label
        first, second = (src_label, tgt_label) if direction == "source_to_target" else (tgt_label, src_label)
        rows.append((first, second, e.detail or ""))
    rows.sort(key=lambda r: (r[0], r[1]))

    _write_sheet(ws, headers, rows, table_name="TableLineage")


def _column_rows(full_graph: LineageGraph) -> list[tuple[str, str, str, str]]:
    """Physical-table-column-to-physical-table-column rows only, collapsing over
    any intermediate transformation/instance column hops (e.g. Informatica
    Source Qualifier / Expression / Lookup instances)."""
    physical_columns = full_graph.collapse_to_physical_columns()
    rows = []
    for e in physical_columns.edges:
        src_table, src_col = _split_table_column(physical_columns.nodes[e.source].label)
        tgt_table, tgt_col = _split_table_column(physical_columns.nodes[e.target].label)
        rows.append((src_table, src_col, tgt_table, tgt_col))
    return rows


def _split_table_column(label: str) -> tuple[str, str]:
    if "." in label:
        table, col = label.rsplit(".", 1)
        return table, col
    return label, ""


def _write_column_sheet(ws, column_rows: list[tuple[str, str, str, str]], direction: str) -> None:
    if direction == "source_to_target":
        headers = ["Source Table", "Source Column", "Target Table", "Target Column"]
        rows = list(column_rows)
    else:
        headers = ["Target Table", "Target Column", "Source Table", "Source Column"]
        rows = [(tt, tc, st, sc) for st, sc, tt, tc in column_rows]

    rows = sorted(set(rows), key=lambda r: (r[0], r[2], r[1]))
    _write_sheet(ws, headers, rows, table_name="ColumnLineage")


def _write_sheet(ws, headers: list[str], rows: list[tuple], table_name: str) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"

    widths = [len(h) for h in headers]
    for row in rows:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], len(str(val)))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(width + 4, 60)

    if rows:
        last_row = len(rows) + 1
        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{last_row}"
        tbl = Table(displayName=table_name, ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tbl)
