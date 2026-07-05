from pathlib import Path

import openpyxl
import pytest

from lineage_mcp import diagram
from lineage_mcp.analyzer import analyze
from lineage_mcp.excel import to_excel_workbook
from lineage_mcp.sql.parser import parse_sql_lineage

FIXTURES = Path(__file__).parent / "fixtures"


def _expected_sheets(*rest: str) -> list[str]:
    return (["Flow Diagram"] if diagram.is_available() else []) + list(rest)


def _load(name: str) -> str:
    return (FIXTURES / name).read_text(encoding="utf-8")


def test_excel_workbook_has_table_and_column_sheets():
    result = parse_sql_lineage(_load("sample.sql"))
    wb = to_excel_workbook(result.graph, direction="source_to_target")
    assert wb.sheetnames == _expected_sheets("Table Lineage", "Column Lineage")

    table_sheet = wb["Table Lineage"]
    assert [c.value for c in table_sheet[1]] == ["Source Table", "Target Table", "Detail"]
    table_rows = [tuple(c.value for c in row) for row in table_sheet.iter_rows(min_row=2)]
    assert any(r[0] == "raw.customers" and r[1] == "stg.customers" for r in table_rows)

    col_sheet = wb["Column Lineage"]
    assert [c.value for c in col_sheet[1]] == ["Source Table", "Source Column", "Target Table", "Target Column"]
    col_rows = [tuple(c.value for c in row) for row in col_sheet.iter_rows(min_row=2)]
    assert ("raw.orders", "order_id", "mart.customer_orders", "order_id") in col_rows


def test_excel_workbook_target_to_source_reorders_columns_not_values():
    result = parse_sql_lineage(_load("sample.sql"))
    wb = to_excel_workbook(result.graph, direction="target_to_source")

    table_sheet = wb["Table Lineage"]
    assert [c.value for c in table_sheet[1]] == ["Target Table", "Source Table", "Detail"]
    table_rows = [tuple(c.value for c in row) for row in table_sheet.iter_rows(min_row=2)]
    assert any(r[0] == "stg.customers" and r[1] == "raw.customers" for r in table_rows)

    col_sheet = wb["Column Lineage"]
    assert [c.value for c in col_sheet[1]] == ["Target Table", "Target Column", "Source Table", "Source Column"]
    col_rows = [tuple(c.value for c in row) for row in col_sheet.iter_rows(min_row=2)]
    assert ("mart.customer_orders", "order_id", "raw.orders", "order_id") in col_rows


def test_excel_workbook_skips_column_sheet_when_no_column_edges():
    result = parse_sql_lineage("MERGE INTO t.tgt USING t.src ON t.tgt.id = t.src.id WHEN MATCHED THEN UPDATE SET id = t.src.id;")
    wb = to_excel_workbook(result.graph)
    assert wb.sheetnames == _expected_sheets("Table Lineage")


def test_excel_workbook_embeds_diagram_image_when_graphviz_available():
    if not diagram.is_available():
        pytest.skip("Graphviz 'dot' executable not available in this environment")
    result = parse_sql_lineage(_load("sample.sql"))
    wb = to_excel_workbook(result.graph)
    diagram_sheet = wb["Flow Diagram"]
    assert len(diagram_sheet._images) == 1


def test_excel_workbook_skips_diagram_sheet_gracefully_without_graphviz(monkeypatch):
    monkeypatch.setattr("lineage_mcp.excel.render_flow_diagram_png", lambda *a, **k: None)
    result = parse_sql_lineage(_load("sample.sql"))
    wb = to_excel_workbook(result.graph)
    assert "Flow Diagram" not in wb.sheetnames
    assert wb.sheetnames == ["Table Lineage", "Column Lineage"]


def test_analyze_writes_excel_file_to_output_dir_relative_to_cwd(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    src_dir = tmp_path / "input"
    src_dir.mkdir()
    src = src_dir / "sample.sql"
    src.write_text(_load("sample.sql"), encoding="utf-8")

    result = analyze(file_path=str(src), output_formats=["excel"])
    excel_path = Path(result["excel_path"])
    assert excel_path == tmp_path / "output" / "sample_lineage.xlsx"
    assert excel_path.exists()

    wb = openpyxl.load_workbook(excel_path)
    assert "Table Lineage" in wb.sheetnames


def test_analyze_excel_requires_path_for_inline_content():
    with pytest.raises(ValueError):
        analyze(content="INSERT INTO t2 SELECT a FROM t1;", file_type="sql", output_formats=["excel"])


def test_analyze_excel_respects_explicit_output_path(tmp_path):
    out = tmp_path / "custom_name.xlsx"
    result = analyze(
        content="INSERT INTO t2 SELECT a FROM t1;",
        file_type="sql",
        output_formats=["excel"],
        excel_path=str(out),
    )
    assert result["excel_path"] == str(out)
    assert out.exists()
